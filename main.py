# STANDARD LIBRARY IMPORTS
from collections import Counter
import re
from datetime import datetime

# THIRD PARTY IMPORTS
import win32com.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# CONSTANTS
SUBJECT = "[External] Message meets Alert condition"
OUTPUT_FILE = "PARSED_INTRUSION_EMAILS.xlsx"

# INPUTS
DATE_START = datetime.strptime(input("Start date (YYYY-MM-DD): "), "%Y-%m-%d")
DATE_END   = datetime.strptime(input("End date (YYYY-MM-DD): "), "%Y-%m-%d")

# FUNCTION
def get_emails(subject: str, start: datetime, end: datetime) -> list[dict]:
    """
    Connect to the Alerts shared mailbox in Outlook and retrieve emails
    filtered by exact subject and date range. Parses email body fields
    using regex.

    Args:
        subject (str): Exact subject line to match against.
        start (datetime): Start of the date range (inclusive).
        end (datetime): End of the date range (inclusive).

    Returns:
        list[dict]: Each dict contains 'date', 'srcip', 'srccountry',
                    'proto', 'service', and 'dstip'.
    """
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")

    inbox = None

    for store in namespace.Stores:
        if store.DisplayName == "alerts":
            inbox = store.GetDefaultFolder(6)
            break

    if not inbox:
        raise ValueError("Alerts mailbox not found.")

    filter_str = (f"[Subject] = '{subject}' AND "
                  f"[ReceivedTime] >= '{start.strftime('%m/%d/%Y')}' AND "
                  f"[ReceivedTime] <= '{end.strftime('%m/%d/%Y')}'")

    all_email_items = inbox.Items
    all_email_items.Sort("[ReceivedTime]", True)
    filtered_email_items = all_email_items.Restrict(filter_str)

    results_list = []

    for email in filtered_email_items:
        body = email.Body

        # regex
        ip = re.search(r'srcip=([\d.]+)', body)
        country = re.search(r'srccountry="([^"]+)"', body)
        proto = re.search(r'proto=(\d+)', body)
        service = re.search(r'service="([^"]+)"', body)
        destip = re.search(r'dstip=([\d.]+)', body)

        # errors
        if ip:
            srcip = ip.group(1)
        else:
            srcip = "NOT FOUND"

        if country:
            srccountry = country.group(1)
        else:
            srccountry = "NOT FOUND"

        if proto:
            srcproto = proto.group(1)
        else:
            srcproto = "NOT FOUND"

        if service:
            srcservice = service.group(1)
        else:
            srcservice = "NOT FOUND"

        if destip:
            dstip = destip.group(1)
        else:
            dstip = "NOT FOUND"

        # dict creation
        results_list.append({
            "date"      : email.ReceivedTime.strftime("%Y-%m-%d %H:%M"),
            "srcip"     : srcip,
            "srccountry": srccountry,
            "proto" : srcproto,
            "service" : srcservice,
            "dstip" : dstip,})

    return results_list

def write_excel(result_list: list[dict], output_path: str):

    # Create a workbook and sheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "PARSED EMAILS"
    # Write styled headers
    headers = ["DATE", "SRCIP", "SRCCOUNTRY", "PROTO", "SERVICE", "DSTIP"]
    header_font = Font(name="Times New Roman", size = 14, bold=True, italic=False, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F4F7F")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True,)

    for column, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Sort data by IP frequency, NOT FOUND at the bottom
    ip_counts = Counter(email["srcip"] for email in result_list if email["srcip"] != "NOT FOUND")

    found = [email for email in result_list if email["srcip"] != "NOT FOUND"]
    not_found = [email for email in result_list if email["srcip"] == "NOT FOUND"]

    found.sort(key=lambda email: ip_counts[email["srcip"]], reverse=True)

    combined_ip = found + not_found
    # Write each row
    not_found_fill = PatternFill(fill_type="solid", fgColor="FFE4E1")

    for row, entry in enumerate(combined_ip, 2):
        worksheet.cell(row=row, column=1, value=entry["date"])
        worksheet.cell(row=row, column=2, value=entry["srcip"])
        worksheet.cell(row=row, column=3, value=entry["srccountry"])
        worksheet.cell(row=row, column=4, value=entry["proto"])
        worksheet.cell(row=row, column=5, value=entry["service"])
        worksheet.cell(row=row, column=6, value=entry["dstip"])

        count = ip_counts.get(entry["srcip"], 0)

        worksheet.cell(row=row, column=7, value=count if count else "")

        if entry["srcip"] == "NOT FOUND":
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).fill = not_found_fill
    # Set column widths
    column_widths = {"A": 20, "B": 18, "C": 25, "D": 15, "E": 18, "F": 10, "G": 15}
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width
    # Save the file
    workbook.save(output_path)
    print(f"Done. {len(result_list)} rows written to {output_path}")

if __name__ == "__main__":
    data = get_emails(SUBJECT, DATE_START, DATE_END)
    write_excel(data, OUTPUT_FILE)